import openpyxl, re, uuid, sys, datetime, os

XLSX = sys.argv[1]
OUT  = sys.argv[2]
EXISTING = set(s for s in open('/tmp/existing_slugs.txt').read().split('|') if s)
NS = uuid.UUID('6f1d2c3e-0000-4000-8000-000000000001')  # same namespace as PDF batch
TODAY = '2026-06-16'

PT = {'single family':'single_family','townhouse':'townhouse','condo':'condo',
      'condominium':'condo','high rise':'condo','mid rise':'condo','low rise':'condo',
      'apartment':'condo'}
STATUS = {'active':'selling','coming soon':'coming_soon','sold out':'sold_out',
          'registration':'coming_soon','selling':'selling','sold':'sold_out'}
CSTATUS = {'pre-construction':'preconstruction','preconstruction':'preconstruction',
           'under construction':'under_construction','construction':'under_construction',
           'completed':'completed','complete':'completed','occupancy':'completed'}
PTLABEL = {'single_family':'single-family homes','townhouse':'townhomes','condo':'condominiums'}

def s(v):
    if v is None: return None
    if isinstance(v, float) and v.is_integer(): v = int(v)
    v = str(v).strip()
    return v or None
def i(v):
    if v is None: return None
    try:
        n = int(float(v));  return n
    except: return None
def money(v):
    if v is None: return None
    try: return round(float(v),2)
    except: return None
def d(v):
    if isinstance(v,(datetime.datetime,datetime.date)): return v.strftime('%Y-%m-%d')
    return None
def mY(v):
    if isinstance(v,(datetime.datetime,datetime.date)): return v.strftime('%b %Y')
    return None
def Q(v): return 'null' if v is None else "'"+str(v).replace("'","''")+"'"
def NQ(v): return 'null' if v in (None,'') else str(v)
def slugify(x):
    x=re.sub(r"[''`]","",x.lower()); x=re.sub(r'[^a-z0-9]+','-',x).strip('-'); return re.sub(r'-+','-',x)

wb = openpyxl.load_workbook(XLSX, data_only=True)
# Optional 3rd arg: sheet name or 0-based index (defaults to first sheet).
if len(sys.argv) > 3:
    sel = sys.argv[3]
    ws = wb[sel] if sel in wb.sheetnames else wb.worksheets[int(sel)]
else:
    ws = wb.worksheets[0]
rows = list(ws.iter_rows(values_only=True))
hdr = [s(h) for h in rows[0]]
# Some exports have a leading blank index column in the DATA rows that the
# header lacks (data shifted one column right). Detect by checking whether the
# 'Inventory number' values actually sit one column right of the header, and
# realign the header if so.
inv_pos = next((n for n, h in enumerate(hdr) if h == 'Inventory number'), None)
if inv_pos is not None:
    sample = [r for r in rows[1:8] if any(c is not None for c in r)]
    here = sum(1 for r in sample if inv_pos < len(r) and r[inv_pos] not in (None, ''))
    right = sum(1 for r in sample if inv_pos + 1 < len(r) and r[inv_pos + 1] not in (None, ''))
    if right > here:
        hdr = [None] + hdr
idx = {h:n for n,h in enumerate(hdr) if h}
def g(row, name): 
    n = idx.get(name); 
    return row[n] if n is not None and n < len(row) else None

recs=[]
for row in rows[1:]:
    inv = s(g(row,'Inventory number'))
    name = s(g(row,'Development name'))
    if not inv or not name: continue
    recs.append({
      'inv':inv,'name':name,
      'status':s(g(row,'Status')),'builder':s(g(row,'Developer')),
      'submarket':s(g(row,'Submarket')),'ptype':s(g(row,'Product type')),
      'remaining':i(g(row,'Current remaining inventory')),'total':i(g(row,'Total units')),
      'market':s(g(row,'Market')),'released':i(g(row,'Total number of units released')),
      'address':s(g(row,'Address')),'cstatus':s(g(row,'Construction status')),
      'occ':g(row,'First occ. date'),'yet':i(g(row,'Units yet to be released')),
      'smin':i(g(row,'Min size (sq. ft)')),'smax':i(g(row,'Max size (sq. ft)')),
      'pmin':money(g(row,'Min price')),'pmax':money(g(row,'Max price')),
      'storeys':i(g(row,'Number of storeys')),'ctype':s(g(row,'Construction type')),
      'maint':s(g(row,'Maintenance fee PPSF')),'lottype':s(g(row,'Lot type')),
      'lotsize':s(g(row,'Lot size (ft.)')),'salesstart':g(row,'Sales start date'),
      'tenure':s(g(row,'Tenure')),'salesteam':s(g(row,'Sales team')),
      'updated':g(row,'Last updated date'),'mortgage':s(g(row,'Mortgage program')),
      'phone':s(g(row,'Contact number')),'coop':s(g(row,'Cooperation status')),
      'mls':s(g(row,'MLS Zone')),'region':s(g(row,'Region')),
      'muni':s(g(row,'Municipality')),'psub':s(g(row,'Product subtype')),
    })

# diagnostics + dedupe by inventory number (keep most-recently-updated)
total_data = sum(1 for row in rows[1:] if any(c is not None for c in row))
print(f"[diag] data rows={total_data} parsed recs={len(recs)}")
by_inv = {}
for r in recs:
    k = r['inv']
    if k not in by_inv:
        by_inv[k] = r
    else:
        a = by_inv[k].get('updated'); b = r.get('updated')
        # keep the row with the later 'Last updated date' (fallback: keep existing)
        if b is not None and (a is None or (hasattr(b,'date') and hasattr(a,'date') and b > a)):
            by_inv[k] = r
recs = list(by_inv.values())
print(f"[diag] distinct inventory={len(recs)} (deduped {len(by_inv)} keys)")

# slugs
used=set(EXISTING)
for r in recs:
    base=slugify(r['name']) or ('project-'+str(r['inv']))
    muni=slugify(r['muni'] or '')
    cand=base
    if cand in used and muni: cand=f"{base}-{muni}"
    n=2
    while cand in used: cand=f"{base}-{n}"; n+=1
    used.add(cand); r['slug']=cand
    r['id']=str(uuid.uuid5(NS,'altus:'+str(r['inv'])))

def own(t):
    if not t: return None
    t=t.lower()
    if 'condominium' in t: return 'condominium'
    if 'common element' in t: return 'freehold_common_element'
    if 'freehold' in t: return 'freehold'
    return None
def descr(r):
    sub=(r['psub']+' ' if r['psub'] else '')
    pt=PTLABEL.get(PT.get((r['ptype'] or '').lower(),''), (r['ptype'] or 'New homes').lower())
    by=f" by {r['builder']}" if r['builder'] else ''
    loc=r['muni'] or 'the GTA'; reg=f" ({r['region']} Region)" if r['region'] else ''
    out=f"{sub}{pt}{by} in {loc}{reg}."
    return out[0].upper()+out[1:]
def notes(r):
    bits=[f"Source: Altus Group new-homes export.",
          f"Altus inventory #{r['inv']}.",
          f"Submarket: {r['submarket']}." if r['submarket'] else None,
          f"Market: {r['market']}." if r['market'] else None,
          f"Remaining inventory: {r['remaining']}." if r['remaining'] is not None else None,
          f"Units released: {r['released']}." if r['released'] is not None else None,
          f"Units yet to be released: {r['yet']}." if r['yet'] is not None else None,
          f"Construction type: {r['ctype']}." if r['ctype'] else None,
          f"Maintenance fee PPSF: {r['maint']}." if r['maint'] else None,
          f"Lot: {r['lottype']} {r['lotsize']}ft.".strip() if (r['lottype'] or r['lotsize']) else None,
          f"Sales start: {d(r['salesstart'])}." if d(r['salesstart']) else None,
          f"Mortgage program: {r['mortgage']}." if r['mortgage'] else None,
          f"MLS Zone: {r['mls']}." if r['mls'] else None,
          f"Sales team: {r['salesteam']}." if r['salesteam'] else None,
          f"Cooperation: {r['coop']}." if r['coop'] else None,
          f"Altus last updated: {d(r['updated'])}." if d(r['updated']) else None,
          f"Imported {TODAY}."]
    return ' '.join(b for b in bits if b)

proj=[]; comm=[]
for r in recs:
    pt=PT.get((r['ptype'] or '').lower())
    ss=STATUS.get((r['status'] or '').lower(),'unknown')
    cs=None
    _cv=(r['cstatus'] or '').lower()
    if 'under construction' in _cv: cs='under_construction'
    elif 'pre-construction' in _cv or 'preconstruction' in _cv: cs='preconstruction'
    elif 'complete' in _cv or 'occupancy' in _cv: cs='completed'
    occ=mY(r['occ'])
    proj.append(f"""('{r['id']}', {Q(r['slug'])}, {Q(r['name'])}, {Q(r['builder'])},
  {Q(pt)}, {Q(ss)}, {Q(cs)}, {Q(own(r['tenure']))},
  {Q(r['address'])}, {Q(r['muni'])}, {Q(r['muni'])}, 'Ontario',
  {NQ(r['total'] if r['total'] else None)}, {NQ(r['storeys'])},
  {NQ(r['smin'])}, {NQ(r['smax'])}, {NQ(r['pmin'])}, {NQ(r['pmax'])}, 'CAD',
  {Q('Est. occupancy '+occ if occ else None)}, {Q(d(r['occ']))},
  {Q(r['phone'])}, {Q(descr(r))},
  'Altus Group', {Q(r['builder'])}, {Q(notes(r))},
  {Q(d(r['updated']))}, true, 'draft', false)""")
    if r['coop'] or r['salesteam']:
        cid=str(uuid.uuid5(NS,'altus-comm:'+str(r['inv'])))
        neg=' '.join(x for x in [f"Cooperation status: {r['coop']}." if r['coop'] else None,
                                 f"Sales team: {r['salesteam']}." if r['salesteam'] else None] if x)
        comm.append(f"('{cid}', '{r['id']}', null, null, {Q(neg)})")

BATCH = 40  # rows per INSERT statement (keeps each statement small enough to apply)
PCOLS = """insert into public.projects (
  id, slug, project_name, builder_name,
  project_type, sales_status, construction_status, ownership_type,
  address_full, city, municipality, province,
  total_units, storeys,
  size_range_sqft_min, size_range_sqft_max, price_from_public, price_to_public, price_currency,
  occupancy_estimate_text, occupancy_start_date,
  sales_centre_phone, description_short,
  external_source, builder_names_raw, import_notes,
  last_verified_at, is_seeded, record_status, public_page_enabled
) values"""
PCONFLICT = """on conflict (id) do update set
  project_name=excluded.project_name, builder_name=excluded.builder_name,
  project_type=excluded.project_type, sales_status=excluded.sales_status,
  construction_status=excluded.construction_status, ownership_type=excluded.ownership_type,
  address_full=excluded.address_full, city=excluded.city, municipality=excluded.municipality,
  province=excluded.province, total_units=excluded.total_units, storeys=excluded.storeys,
  size_range_sqft_min=excluded.size_range_sqft_min, size_range_sqft_max=excluded.size_range_sqft_max,
  price_from_public=excluded.price_from_public, price_to_public=excluded.price_to_public,
  price_currency=excluded.price_currency, occupancy_estimate_text=excluded.occupancy_estimate_text,
  occupancy_start_date=excluded.occupancy_start_date, sales_centre_phone=excluded.sales_centre_phone,
  description_short=excluded.description_short, external_source=excluded.external_source,
  builder_names_raw=excluded.builder_names_raw, import_notes=excluded.import_notes,
  last_verified_at=excluded.last_verified_at, updated_at=now();"""
CCOLS = """insert into public.project_private_commercials (
  id, project_id, commission_summary, commission_percent, negotiability_notes
) values"""
CCONFLICT = """on conflict (project_id) do update set
  commission_summary=excluded.commission_summary, commission_percent=excluded.commission_percent,
  negotiability_notes=excluded.negotiability_notes, updated_at=now();"""

sql=[]
sql.append("-- LIQWD — Altus Group batch import")
sql.append(f"-- Generated {TODAY} from structured Altus xlsx export. {len(proj)} projects.")
sql.append("-- Idempotent: deterministic uuid5 ids + ON CONFLICT. Re-runnable. Batched INSERTs.")
sql.append("-- Altus inventory numbers live ONLY in admin-only provenance (import_notes); never public.\n")
for k in range(0, len(proj), BATCH):
    sql.append(PCOLS)
    sql.append(",\n".join(proj[k:k+BATCH]))
    sql.append(PCONFLICT + "\n")
for k in range(0, len(comm), BATCH):
    sql.append(CCOLS)
    sql.append(",\n".join(comm[k:k+BATCH]))
    sql.append(CCONFLICT + "\n")
open(OUT,'w').write("\n".join(sql))
print(f"projects={len(proj)} commercials={len(comm)} -> {OUT}")
print("construction_status values seen:", sorted(set((r['cstatus'] or '') for r in recs)))
print("sample slugs:", [r['slug'] for r in recs[:6]])
